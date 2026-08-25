"""Управление пользователями (админ)."""
import csv
import io
import secrets

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .. import models, schemas
from ..database import get_db
from ..deps import require_admin
from ..security import hash_password

router = APIRouter(prefix="/api/users", tags=["users"])

# Синонимы заголовков колонок (рус/англ) → каноническое имя поля.
_HEADER_ALIASES = {
    "email": "email", "e-mail": "email", "почта": "email", "мейл": "email",
    "full_name": "full_name", "fio": "full_name", "фио": "full_name",
    "имя": "full_name", "name": "full_name", "сотрудник": "full_name",
    "department": "department", "отдел": "department",
    "подразделение": "department",
    "position": "position", "должность": "position",
    "role": "role", "роль": "role",
    "password": "password", "пароль": "password",
}
_ROLE_ALIASES = {
    "admin": "admin", "администратор": "admin", "админ": "admin",
    "teacher": "teacher", "преподаватель": "teacher", "учитель": "teacher",
    "student": "student", "студент": "student", "сотрудник": "student",
}


def _parse_rows(filename: str, data: bytes) -> list[dict]:
    """Разобрать CSV или XLSX в список словарей {заголовок: значение}."""
    name = (filename or "").lower()
    rows: list[dict] = []
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h or "").strip().lower() for h in (next(it, []) or [])]
        for r in it:
            rows.append(
                {header[i]: r[i] for i in range(min(len(header), len(r)))}
            )
    else:
        text = data.decode("utf-8-sig", errors="replace")
        delimiter = ";" if text.count(";") > text.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        for r in reader:
            rows.append({(k or "").strip().lower(): v for k, v in r.items()})
    return rows


@router.get("", response_model=list[schemas.UserOut])
async def list_users(
    q: str | None = None,
    role: models.UserRole | None = None,
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: AsyncSession = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    stmt = select(models.User).order_by(models.User.created_at.desc())
    if role:
        stmt = stmt.where(models.User.role == role)
    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            models.User.full_name.ilike(like) | models.User.email.ilike(like)
        )
    stmt = stmt.limit(limit).offset(offset)
    result = await db.execute(stmt)
    return result.scalars().all()



@router.post("/import", response_model=schemas.UserImportResult)
async def import_users(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    """Массовое создание пользователей из CSV/XLSX (как в Moodle)."""
    raw = await file.read()
    try:
        rows = _parse_rows(file.filename or "", raw)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {exc}")

    created = 0
    skipped = 0
    accounts: list[schemas.ImportedAccount] = []
    issues: list[schemas.ImportRowIssue] = []
    seen: set[str] = set()

    for idx, row in enumerate(rows, start=2):  # строка 1 — заголовки
        norm: dict[str, str] = {}
        for key, value in row.items():
            canon = _HEADER_ALIASES.get((key or "").strip().lower())
            if canon:
                norm[canon] = "" if value is None else str(value).strip()

        email = norm.get("email", "").strip().lower()
        full_name = norm.get("full_name", "").strip()
        if not email and not full_name:
            continue  # пустая строка

        if not email or "@" not in email:
            issues.append(schemas.ImportRowIssue(
                row=idx, email=email, reason="Некорректный email"))
            continue
        if not full_name:
            issues.append(schemas.ImportRowIssue(
                row=idx, email=email, reason="Пустое ФИО"))
            continue
        if email in seen:
            skipped += 1
            issues.append(schemas.ImportRowIssue(
                row=idx, email=email, reason="Дубликат в файле"))
            continue
        exists = await db.scalar(
            select(models.User).where(models.User.email == email)
        )
        if exists:
            skipped += 1
            issues.append(schemas.ImportRowIssue(
                row=idx, email=email, reason="Email уже существует"))
            continue

        role_key = _ROLE_ALIASES.get(norm.get("role", "").strip().lower(), "student")
        password = norm.get("password", "").strip() or secrets.token_urlsafe(6)
        db.add(models.User(
            email=email,
            full_name=full_name,
            department=norm.get("department") or None,
            position=norm.get("position") or None,
            role=models.UserRole(role_key),
            hashed_password=hash_password(password),
        ))
        seen.add(email)
        created += 1
        accounts.append(schemas.ImportedAccount(
            email=email, full_name=full_name, password=password))

    if created:
        await db.commit()
    return schemas.UserImportResult(
        created=created, skipped=skipped, accounts=accounts, issues=issues)

@router.post("", response_model=schemas.UserOut, status_code=201)
async def create_user(
    data: schemas.UserCreate,
    db: AsyncSession = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    exists = await db.scalar(select(models.User).where(models.User.email == data.email))
    if exists:
        raise HTTPException(status_code=400, detail="Email уже занят")
    user = models.User(
        email=data.email,
        full_name=data.full_name,
        department=data.department,
        position=data.position,
        role=data.role,
        hashed_password=hash_password(data.password),
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


@router.patch("/{user_id}", response_model=schemas.UserOut)
async def update_user(
    user_id: int,
    data: schemas.UserUpdate,
    db: AsyncSession = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    user = await db.get(models.User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    payload = data.model_dump(exclude_unset=True)
    if "password" in payload and payload["password"]:
        user.hashed_password = hash_password(payload.pop("password"))
    else:
        payload.pop("password", None)
    for field, value in payload.items():
        setattr(user, field, value)
    await db.commit()
    await db.refresh(user)
    return user


@router.delete("/{user_id}", status_code=204)
async def delete_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    if user_id == admin.id:
        raise HTTPException(status_code=400, detail="Нельзя удалить самого себя")
    user = await db.get(models.User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    await db.delete(user)
    await db.commit()
